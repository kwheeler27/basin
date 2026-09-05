"""Bake Upper Basin consumptive-use history from Reclamation's CU&L workbook.

Source of record: the Upper Colorado River System Consumptive Uses and
Losses (CU&L) data-summary workbook published on the UC Region document
library (usbr.gov/uc/DocLibrary/reports.html). Unlike the Lower Basin's
decree accounting, these are MODELED ESTIMATES (the CU&L methods manual
covers 1991-2025), revised in versioned workbook releases — the artifact
records the release version and the workbook's own title.

Extracted per calendar year: the Upper Basin total consumptive use
including undistributed mainstem-reservoir evaporation (the CU&L headline
number), the five-state distributed total excluding that evaporation (the
other federal bucketing of the same estimate), the evaporation itself,
and each state's total. Every year is validated:
  - the five states must sum to the distributed total within 1%
  - distributed total + reservoir evaporation must equal the headline
    total within 1%
  - every value must sit in a plausible historical range
A year that fails validation is EXCLUDED and reported — never silently
guessed. (The current workbook carries a placeholder column for a year
its own title excludes; the range gate catches it.)

Requires: openpyxl.
Run from packages/ingest:  python3 bake_ub_consumption.py
Cadence: annual-ish, when Reclamation posts a new workbook release
(check the reports page for a new filename — the URL embeds the version).
"""

from __future__ import annotations

import datetime as dt
import io
import json
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

URL = (
    "https://www.usbr.gov/uc/DocLibrary/Reports/ConsumptiveUsesLosses/"
    "20251201-UpperColoradoCUL1971-2025_v24.5_MajorTribSummary-DataSummariesFigures.xlsx"
)
VERSION = "v24.5"
OUT = (
    Path(__file__).resolve().parents[2]
    / "apps" / "web" / "public" / "geo" / "ub_consumption_cy.json"
)
UA = "Mozilla/5.0 basin/0.1"

STATES = {
    "Arizona": "az",
    "Colorado": "co",
    "New Mexico": "nm",
    "Utah": "ut",
    "Wyoming": "wy",
}

# Plausible ranges (acre-feet) — guards against grabbing a wrong number
# and against placeholder columns for years the workbook doesn't cover.
RANGES = {
    "ubTotal": (2_500_000, 6_500_000),
    "statesTotal": (2_200_000, 6_000_000),
    "resEvap": (80_000, 1_000_000),
    "az": (1_000, 80_000),
    "co": (1_100_000, 2_700_000),
    "nm": (140_000, 600_000),
    "ut": (300_000, 1_400_000),
    "wy": (150_000, 900_000),
}


def fetch_workbook() -> tuple[bytes, str]:
    req = urllib.request.Request(URL, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=180) as r:
        return r.read(), dt.date.today().isoformat()


def label_of(row: tuple) -> str:
    """The row's text label: every string cell left of the year columns."""
    return " ".join(str(c).strip() for c in row[:6] if isinstance(c, str) and c.strip())


def parse(blob: bytes) -> tuple[dict, dict[str, dict[str, int]]]:
    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb["Summary"]
    rows = list(ws.iter_rows(values_only=True))

    title = next(
        (str(c) for r in rows[:3] for c in r if isinstance(c, str) and "Consumptive" in c),
        "",
    )

    header = next(r for r in rows[:8] if any(c == "USE TYPE" for c in r))
    year_cols = {int(c): i for i, c in enumerate(header) if isinstance(c, (int, float))}

    # The first block of the sheet runs TOTAL CONSUMPTIVE USE → per-state
    # rows → Grand Total (excl. reservoir evap) → undistributed reservoir
    # evaporation → subtotal → back-check. Rows are located by label, not
    # index, so a workbook revision that inserts a row doesn't misparse.
    ub_total = states_total = res_evap = subtotal = None
    state_rows: dict[str, list[tuple]] = {}
    current_state: str | None = None
    in_states = False
    for row in rows:
        lab = label_of(row)
        if lab.startswith("TOTAL CONSUMPTIVE USE"):
            ub_total = row
            in_states = True
            continue
        if lab.startswith("Grand Total"):
            states_total = row
            in_states = False
            current_state = None
            continue
        if lab.startswith("CU&L undistributed reservoir evaporation"):
            res_evap = row
            continue
        # Must break at the first block's subtotal: the sheet repeats the
        # same footer labels under every later use-type block.
        if lab.startswith("subtotal"):
            subtotal = row
            break
        if in_states:
            for name in STATES:
                if any(c == name for c in row[:6]):
                    current_state = name
                    break
            if current_state is not None and any(
                isinstance(c, (int, float)) for c in row[6:]
            ):
                state_rows.setdefault(current_state, []).append(row)

    for what, row in [
        ("TOTAL CONSUMPTIVE USE", ub_total),
        ("Grand Total", states_total),
        ("undistributed reservoir evaporation", res_evap),
        ("subtotal", subtotal),
    ]:
        if row is None:
            raise SystemExit(f"structure changed: '{what}' row not found")

    def state_total_row(name: str) -> tuple:
        cands = state_rows.get(name, [])
        totals = [r for r in cands if any(c == "Total" for c in r[:6])]
        if totals:
            return totals[0]
        if len(cands) == 1:
            return cands[0]
        raise SystemExit(
            f"structure changed: {name} has {len(cands)} rows and no Total row"
        )

    component_rows: dict[str, tuple] = {
        "ubTotal": ub_total,
        "statesTotal": states_total,
        "resEvap": res_evap,
        **{key: state_total_row(name) for name, key in STATES.items()},
    }

    years: dict[str, dict[str, int]] = {}
    excluded: list[int] = []
    for y, col in sorted(year_cols.items()):
        vals: dict[str, float] = {}
        ok = True
        for key, row in component_rows.items():
            v = row[col]
            if not isinstance(v, (int, float)):
                ok = False
                break
            lo, hi = RANGES[key]
            if not (lo <= v <= hi):
                ok = False
                break
            vals[key] = v
        if ok:
            state_sum = sum(vals[k] for k in STATES.values())
            if abs(state_sum - vals["statesTotal"]) > 0.01 * vals["statesTotal"]:
                ok = False
            elif (
                abs(vals["statesTotal"] + vals["resEvap"] - vals["ubTotal"])
                > 0.01 * vals["ubTotal"]
            ):
                ok = False
        if not ok:
            excluded.append(y)
            continue
        years[str(y)] = {k: round(v) for k, v in vals.items()}

    if len(years) < 40:
        raise SystemExit(f"only {len(years)} valid years — refusing to bake")

    return {
        "source": (
            "US Bureau of Reclamation, Upper Colorado River System Consumptive "
            f"Uses and Losses (CU&L) data-summary workbook, release {VERSION} — "
            "modeled estimates, one per calendar year: Upper Basin total "
            "consumptive use with and without undistributed mainstem-reservoir "
            "evaporation, and each state's total"
        ),
        "workbookTitle": title,
        "url": URL,
        "accountingConcept": (
            "consumptive_use (UB states, distributed) · evaporation "
            "(undistributed mainstem reservoirs) · their stated sum (ubTotal)"
        ),
        "measurementClass": "estimated",
        "excludedYears": excluded,
    }, years


def main() -> None:
    blob, fetched = fetch_workbook()
    meta, years = parse(blob)
    out = {**meta, "fetched": fetched, "years": years}
    OUT.write_text(json.dumps(out, indent=1) + "\n")
    ys = sorted(int(y) for y in years)
    last = years[str(ys[-1])]
    print(f"baked {len(ys)} years ({ys[0]}–{ys[-1]}) → {OUT}")
    print(f"excluded (failed validation): {meta['excludedYears']}")
    print(
        f"latest {ys[-1]}: total {last['ubTotal']:,} af "
        f"(states {last['statesTotal']:,} + reservoir evap {last['resEvap']:,})"
    )


if __name__ == "__main__":
    main()
