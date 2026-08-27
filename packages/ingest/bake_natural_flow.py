"""Bake Reclamation's Lees Ferry annual natural flow to a web artifact.

Source of record: the Colorado River Basin Natural Flow database
(usbr.gov/lc/region/g4000/NaturalFlow/), currently posted as
NaturalFlows1906-2020_20221215.xlsx. Sheet AnnualWYTotalNaturalFlow,
column for USGS 09380000 (Colorado River at Lees Ferry) — TOTAL natural
flow at the node, water-year totals in acre-feet.

"Natural" flow is a computed quantity: observed flow with upstream
consumptive use and reservoir operations added back. Reclamation revises
it; the vintage is part of the artifact.

Run from packages/ingest:  python3 bake_natural_flow.py
Cadence: annual — re-run when Reclamation posts a new workbook (the
lookout does not watch this page; check with the decree-accounting
annual refresh).
"""

from __future__ import annotations

import datetime as dt
import json
import tempfile
import urllib.request
from pathlib import Path

import openpyxl

URL = (
    "https://www.usbr.gov/lc/region/g4000/NaturalFlow/"
    "NaturalFlows1906-2020_20221215.xlsx"
)
VINTAGE = "2022-12-15 release, WY1906-2020"
GAUGE = "09380000"
OUT = (
    Path(__file__).resolve().parents[2]
    / "apps" / "web" / "public" / "geo" / "natural_flow_wy.json"
)


def main() -> None:
    req = urllib.request.Request(URL, headers={"User-Agent": "Mozilla/5.0 basin/0.1"})
    with urllib.request.urlopen(req, timeout=120) as r:
        blob = r.read()
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as f:
        f.write(blob)
        f.flush()
        wb = openpyxl.load_workbook(f.name, read_only=True, data_only=True)
    ws = wb["AnnualWYTotalNaturalFlow"]

    # Locate the Lees Ferry TOTAL column by its USGS gauge id (row 3).
    gauge_row = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
    col = next(i for i, g in enumerate(gauge_row) if str(g) == GAUGE)

    wy: dict[int, int] = {}
    for row in ws.iter_rows(min_row=7, values_only=True):
        y, v = row[2], row[col]
        if isinstance(y, (int, float)) and isinstance(v, (int, float)) and 1900 < y < 2100:
            wy[int(y)] = round(v)

    years = sorted(wy)
    assert years[0] == 1906 and len(years) >= 110, "unexpected sheet shape"

    payload = {
        "source": (
            "US Bureau of Reclamation, Colorado River Basin Natural Flow "
            "database — Colorado River at Lees Ferry (USGS 09380000), "
            "water-year TOTAL natural flow, acre-feet"
        ),
        "url": URL,
        "vintage": VINTAGE,
        "accountingConcept": "flow (naturalized — computed, not gauged)",
        "measurementClass": "estimated",
        "fetched": dt.date.today().isoformat(),
        "wy": {str(y): wy[y] for y in years},
    }
    OUT.write_text(json.dumps(payload))
    mean_all = sum(wy.values()) / len(wy)
    recent = [wy[y] for y in years if y >= 2000]
    print(
        f"wrote {OUT.name}: WY{years[0]}-{years[-1]} ({len(years)} yrs), "
        f"mean {mean_all/1e6:.2f} MAF, 2000+ mean {sum(recent)/len(recent)/1e6:.2f} MAF"
    )


if __name__ == "__main__":
    main()
